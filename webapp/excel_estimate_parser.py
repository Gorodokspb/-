"""Parse Excel estimate files into section/item rows for standalone estimates.

Uses openpyxl (read-only) to extract rows from .xlsx files and converts them
into `ParsedEstimateRow` objects compatible with `EstimateItemInput`.
No database access, no route dependencies — pure parsing.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Sequence

import openpyxl
from openpyxl.utils import get_column_letter

MAX_XLSX_BYTES = 2 * 1024 * 1024
MAX_PARSED_ROWS = 500
HEADER_SCAN_ROWS = 10

COLUMN_DEFINITIONS = {
    "name": {
        "canonical": "name",
        "aliases": (
            "наименование", "наименование работ", "наименование услуг",
            "вид работ", "вид услуг", "работы", "работа", "услуги",
            "название", "name", "описание", "description",
            "строительные работы", "позиция",
        ),
    },
    "unit": {
        "canonical": "unit",
        "aliases": (
            "ед.", "ед. изм.", "ед.изм.", "ед изм", "единица",
            "единица измерения", "единицы", "unit", "uom", "мера",
            "ед.измерения", "ед измерения",
        ),
    },
    "quantity": {
        "canonical": "quantity",
        "aliases": (
            "кол-во", "количество", "объем", "объём",
            "обьем", "обьём", "quantity", "qty", "кол.",
            "число", "штук", "площадь",
        ),
    },
    "price": {
        "canonical": "price",
        "aliases": (
            "цена", "цена за ед.", "цена за ед", "цена за единицу",
            "цена за м2", "цена за м²", "цена за м.кв.", "цена за м.кв",
            "стоимость ед.", "стоимость единицы", "price", "расценка",
            "единичная расценка", "тариф",
        ),
    },
    "total": {
        "canonical": "total",
        "aliases": (
            "сумма", "стоимость", "итого", "всего", "общая стоимость",
            "total", "amount", "цена всего", "цена итого",
            "стоимость работ", "итоговая стоимость",
        ),
    },
}

DEFAULT_COLUMN_MAPPING = {
    "name": "A",
    "unit": "B",
    "quantity": "C",
    "price": "D",
    "total": "E",
}


@dataclass(frozen=True)
class ColumnMapping:
    """Maps logical column keys to Excel column letters."""

    name: str = "A"
    unit: str | None = None
    quantity: str | None = None
    price: str | None = None
    total: str | None = None


@dataclass(frozen=True)
class ParsedEstimateRow:
    """Single parsed row from an Excel estimate."""

    row_type: str
    name: str
    sort_order: int
    unit: str | None = None
    quantity: Decimal | None = None
    price: Decimal | None = None
    total: Decimal | None = None
    discounted_total: Decimal | None = None
    reference: str | None = None

    @property
    def is_section(self) -> bool:
        return self.row_type == "section"

    @property
    def is_item(self) -> bool:
        return self.row_type == "item"


@dataclass(frozen=True)
class ExcelEstimateParseResult:
    """Result of parsing an Excel estimate file."""

    rows: list[ParsedEstimateRow] = field(default_factory=list)
    section_count: int = 0
    item_count: int = 0
    parsed_sheet_name: str = ""
    total_rows_in_sheet: int = 0
    diagnostics: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return len(self.rows) > 0


class ExcelEstimateParseError(ValueError):
    """Raised when an Excel estimate file cannot be parsed."""


def _normalize_text(value) -> str:
    if value is None:
        return ""
    try:
        if value != value:
            return ""
    except TypeError:
        pass
    return str(value).strip()


def parse_decimal(value) -> Decimal | None:
    """Parse a numeric value to Decimal, handling commas and whitespace."""
    if value is None:
        return None
    try:
        if value != value:
            return None
    except TypeError:
        pass
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None


def _header_contains_keywords(row_values: list[str]) -> bool:
    """Check if a row contains known header keywords."""
    row_set = {v.casefold() for v in row_values if v}
    total_aliases = 0
    matched = 0
    for col_def in COLUMN_DEFINITIONS.values():
        total_aliases += 1
        for alias in col_def["aliases"]:
            if alias.casefold() in row_set:
                matched += 1
                break
    return matched >= 2


def resolve_estimate_columns(headers: list[str]) -> ColumnMapping:
    """Map Excel header strings to logical column keys.

    Scans row_values for known column aliases and returns a ColumnMapping
    with Excel column letters. Falls back to DEFAULT_COLUMN_MAPPING if
    fewer than 2 columns are recognized.
    """
    available: dict[str, str] = {}
    for column_idx, header_value in enumerate(headers):
        key = _normalize_text(header_value).casefold()
        if key:
            available[key] = get_column_letter(column_idx + 1)

    resolved: dict[str, str | None] = {}
    recognized = 0
    for logical_key, col_def in COLUMN_DEFINITIONS.items():
        found = False
        for alias in col_def["aliases"]:
            if alias.casefold() in available:
                resolved[logical_key] = available[alias.casefold()]
                recognized += 1
                found = True
                break
        if not found:
            resolved[logical_key] = None

    if recognized < 2:
        return ColumnMapping(**DEFAULT_COLUMN_MAPPING)

    return ColumnMapping(
        name=resolved.get("name") or "A",
        unit=resolved.get("unit"),
        quantity=resolved.get("quantity"),
        price=resolved.get("price"),
        total=resolved.get("total"),
    )


def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Scan the first HEADER_SCAN_ROWS for a header row containing column keywords.

    Returns the 1-based row number of the header. Returns 0 if no header recognized,
    meaning parsing starts from the first data row.
    """
    for row_idx in range(1, min(ws.max_row or 1, HEADER_SCAN_ROWS) + 1):
        row_values: list[str] = []
        for cell in ws[row_idx]:
            row_values.append(_normalize_text(cell.value))
        if _header_contains_keywords(row_values):
            return row_idx
    return 0


def _cell_value_in_merged(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int) -> str:
    """Return the value of a cell, resolving merged cells to the top-left value."""
    col_letter = get_column_letter(col)
    cell = ws[f"{col_letter}{row}"]
    val = _normalize_text(cell.value)
    if val:
        return val
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = merged_range.min_col
            top_row = merged_range.min_row
            return _normalize_text(ws.cell(row=top_row, column=top_left).value)
    return ""


def _read_row_values(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    mapping: ColumnMapping,
) -> dict[str, str | None]:
    """Read named column values from a row using the column mapping."""
    result: dict[str, str | None] = {}
    for logical_key, col_letter in (
        ("name", mapping.name),
        ("unit", mapping.unit),
        ("quantity", mapping.quantity),
        ("price", mapping.price),
        ("total", mapping.total),
    ):
        if col_letter is None:
            result[logical_key] = None
            continue
        col_num = openpyxl.utils.column_index_from_string(col_letter)
        result[logical_key] = _cell_value_in_merged(ws, row_idx, col_num)
    return result


def _row_is_empty(row_values: dict[str, str | None]) -> bool:
    """Check if all mapped fields are empty."""
    return all(not (v or "").strip() for v in row_values.values() if v is not None)


def _looks_like_section(row_values: dict[str, str | None]) -> bool:
    """Heuristic: a section row has a name but no numeric quantity/price/total."""
    name = (row_values.get("name") or "").strip()
    if not name:
        return False
    qty = parse_decimal(row_values.get("quantity"))
    price = parse_decimal(row_values.get("price"))
    total = parse_decimal(row_values.get("total"))
    unit = (row_values.get("unit") or "").strip()
    if qty is not None or price is not None or total is not None:
        return False
    name_lower = name.casefold()
    unit_lower = unit.casefold()
    if unit == "" or unit is None or unit_lower == name_lower:
        return True
    return len(unit) > 20


def _looks_like_item(row_values: dict[str, str | None]) -> bool:
    """Heuristic: an item row has a name and at least one numeric or unit field."""
    name = (row_values.get("name") or "").strip()
    if not name:
        return False
    qty = parse_decimal(row_values.get("quantity"))
    price = parse_decimal(row_values.get("price"))
    total = parse_decimal(row_values.get("total"))
    unit = (row_values.get("unit") or "").strip()
    return qty is not None or price is not None or total is not None or unit != ""


def _compute_total(quantity: Decimal | None, price: Decimal | None, total: Decimal | None) -> Decimal | None:
    """Return total from Excel if present, otherwise compute quantity * price."""
    if total is not None and total != Decimal("0"):
        return total
    if quantity is not None and price is not None and quantity != Decimal("0"):
        return quantity * price
    return total


def parse_estimate_xlsx(
    file_bytes: bytes,
    *,
    sheet_name: str | None = None,
) -> ExcelEstimateParseResult:
    """Parse an .xlsx estimate file into a list of ParsedEstimateRow.

    Args:
        file_bytes: Raw bytes of the .xlsx file.
        sheet_name: Specific sheet to parse. Uses first sheet if None.

    Returns:
        ExcelEstimateParseResult with parsed rows and diagnostics.

    Raises:
        ExcelEstimateParseError: If the file cannot be opened or has no
            parseable rows.
    """
    if len(file_bytes) > MAX_XLSX_BYTES:
        raise ExcelEstimateParseError(
            f"Файл слишком большой: {len(file_bytes)} байт. "
            f"Максимум: {MAX_XLSX_BYTES} байт (2 МБ)."
        )

    if len(file_bytes) == 0:
        raise ExcelEstimateParseError("Файл пустой.")

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
        )
    except Exception as exc:
        raise ExcelEstimateParseError(
            f"Не удалось открыть Excel-файл: {exc}"
        ) from exc

    target_ws: openpyxl.worksheet.worksheet.Worksheet | None = None
    if sheet_name:
        if sheet_name in wb.sheetnames:
            target_ws = wb[sheet_name]
        else:
            wb.close()
            raise ExcelEstimateParseError(
                f"Лист '{sheet_name}' не найден. Доступные листы: {', '.join(wb.sheetnames)}"
            )
    elif wb.sheetnames:
        target_ws = wb[wb.sheetnames[0]]
    else:
        wb.close()
        raise ExcelEstimateParseError("В Excel-файле нет листов.")

    header_row_idx = _find_header_row(target_ws)
    if header_row_idx > 0:
        header_values = [_normalize_text(cell.value) for cell in target_ws[header_row_idx]]
        column_mapping = resolve_estimate_columns(header_values)
        data_start_row = header_row_idx + 1
    else:
        column_mapping = resolve_estimate_columns([])
        data_start_row = 1

    diagnostics: list[str] = []
    parsed_rows: list[ParsedEstimateRow] = []
    section_count = 0
    item_count = 0

    sort_order = 0
    max_scanned = min(target_ws.max_row or 1, data_start_row - 1 + MAX_PARSED_ROWS)

    for row_idx in range(data_start_row, max_scanned + 1):
        row_values = _read_row_values(target_ws, row_idx, column_mapping)

        if _row_is_empty(row_values):
            continue

        if _looks_like_section(row_values):
            sort_order += 1
            section_count += 1
            parsed_rows.append(ParsedEstimateRow(
                row_type="section",
                name=row_values["name"].strip(),
                sort_order=sort_order,
            ))
            continue

        if _looks_like_item(row_values):
            sort_order += 1
            item_count += 1
            quantity = parse_decimal(row_values.get("quantity"))
            price = parse_decimal(row_values.get("price"))
            total_raw = parse_decimal(row_values.get("total"))
            total = _compute_total(quantity, price, total_raw)
            discounted_total = total

            unit = (row_values.get("unit") or "").strip() or None

            parsed_rows.append(ParsedEstimateRow(
                row_type="item",
                name=row_values["name"].strip(),
                sort_order=sort_order,
                unit=unit,
                quantity=quantity,
                price=price,
                total=total,
                discounted_total=discounted_total,
                reference=None,
            ))
            continue

        diagnostics.append(
            f"Строка {row_idx} не распознана: name='{row_values.get('name', '')}'"
        )

    wb.close()

    total_scanned = max_scanned - data_start_row + 1
    if total_scanned >= MAX_PARSED_ROWS:
        diagnostics.append(
            f"Достигнут лимит строк ({MAX_PARSED_ROWS}). "
            f"Обработано {item_count + section_count} строк из {total_scanned}."
        )

    if not parsed_rows:
        raise ExcelEstimateParseError(
            "В Excel-файле не найдено строк сметы. "
            "Проверьте, что файл содержит разделы (с названием без чисел) "
            "или позиции работ (с единицами измерения, количеством или ценой)."
        )

    return ExcelEstimateParseResult(
        rows=parsed_rows,
        section_count=section_count,
        item_count=item_count,
        parsed_sheet_name=target_ws.title,
        total_rows_in_sheet=total_scanned,
        diagnostics=diagnostics,
    )


def parsed_rows_to_estimate_items(
    parsed_rows: Sequence[ParsedEstimateRow],
) -> list:
    """Convert ParsedEstimateRow objects to EstimateItemInput-compatible dicts.

    This is a bridge between the parser output and the estimate repository.
    Does NOT import estimate_repository to avoid circular deps.
    """
    items = []
    for row in parsed_rows:
        items.append({
            "name": row.name,
            "sort_order": row.sort_order,
            "row_type": row.row_type,
            "unit": row.unit,
            "quantity": str(row.quantity) if row.quantity is not None else None,
            "price": str(row.price) if row.price is not None else None,
            "total": str(row.total) if row.total is not None else None,
            "discounted_total": str(row.discounted_total) if row.discounted_total is not None else None,
            "parent_section_id": None,
            "section_key": None,
            "reference": row.reference,
            "price_source_type": None,
            "price_source_id": None,
            "is_manual_price": True,
            "notes": None,
        })
    return items
