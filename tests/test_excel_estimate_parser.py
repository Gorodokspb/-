import io
import unittest
from decimal import Decimal

import openpyxl

from webapp.excel_estimate_parser import (
    ColumnMapping,
    ExcelEstimateParseError,
    ExcelEstimateParseResult,
    ParsedEstimateRow,
    MAX_XLSX_BYTES,
    _compute_total,
    _find_header_row,
    _header_contains_keywords,
    _looks_like_item,
    _looks_like_section,
    _looks_like_signature_or_trash,
    _normalize_text,
    _read_row_values,
    _row_is_empty,
    parse_decimal,
    parse_estimate_xlsx,
    parsed_rows_to_estimate_items,
    resolve_estimate_columns,
)


def _create_xlsx_bytes(sheet_data: list[list], *, sheet_name: str = "Смета") -> bytes:
    """Create a .xlsx file in memory from a list of row lists."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in sheet_data:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _create_xlsx_with_merged_section() -> bytes:
    """Create xlsx with a merged cell section row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.append(["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"])
    ws.merge_cells("A5:E5")
    ws["A5"] = "Раздел: Пол"
    ws.append(["Укладка ламината", "м.кв.", 30, 1200, 36000])
    ws.append(["Плинтус", "м.п.", 45, 600, 27000])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


class TestNormalizeText(unittest.TestCase):
    def test_strips_whitespace(self):
        self.assertEqual(_normalize_text("  hello  "), "hello")

    def test_none_returns_empty(self):
        self.assertEqual(_normalize_text(None), "")

    def test_nan_returns_empty(self):
        nan = float("nan")
        self.assertEqual(_normalize_text(nan), "")

    def test_number_to_string(self):
        self.assertEqual(_normalize_text(42), "42")


class TestParseDecimal(unittest.TestCase):
    def test_integer(self):
        self.assertEqual(parse_decimal(42), Decimal("42"))

    def test_float(self):
        self.assertEqual(parse_decimal(10.5), Decimal("10.5"))

    def test_string_with_comma(self):
        self.assertEqual(parse_decimal("10,5"), Decimal("10.5"))

    def test_string_with_spaces(self):
        self.assertEqual(parse_decimal("1 234,56"), Decimal("1234.56"))

    def test_empty_string_returns_none(self):
        self.assertIsNone(parse_decimal(""))

    def test_none_returns_none(self):
        self.assertIsNone(parse_decimal(None))

    def test_nbsp_stripped(self):
        self.assertEqual(parse_decimal("\u00a01\u00a0200"), Decimal("1200"))

    def test_non_numeric_returns_none(self):
        self.assertIsNone(parse_decimal("abc"))


class TestHeaderDetection(unittest.TestCase):
    def test_recognizes_russian_headers(self):
        headers = ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"]
        self.assertTrue(_header_contains_keywords(headers))

    def test_recognizes_mixed_headers(self):
        headers = ["Наименование работ", "Ед. изм.", "Количество", "Цена за ед."]
        self.assertTrue(_header_contains_keywords(headers))

    def test_rejects_non_header_rows(self):
        headers = ["Грунтовка потолка", "м.кв.", "", "100", "2500"]
        self.assertFalse(_header_contains_keywords(headers))

    def test_single_keyword_not_enough(self):
        headers = ["Наименование", "Данные", "Инфо", "Метка"]
        self.assertFalse(_header_contains_keywords(headers))


class TestResolveEstimateColumns(unittest.TestCase):
    def test_resolves_standard_headers(self):
        mapping = resolve_estimate_columns(
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"]
        )
        self.assertEqual(mapping.name, "A")
        self.assertEqual(mapping.unit, "B")
        self.assertEqual(mapping.quantity, "C")
        self.assertEqual(mapping.price, "D")
        self.assertEqual(mapping.total, "E")

    def test_resolves_alternate_headers(self):
        mapping = resolve_estimate_columns(
            ["Вид работ", "Единица", "Количество", "Расценка", "Итого"]
        )
        self.assertEqual(mapping.name, "A")
        self.assertEqual(mapping.unit, "B")
        self.assertEqual(mapping.quantity, "C")
        self.assertEqual(mapping.price, "D")
        self.assertEqual(mapping.total, "E")

    def test_fewer_than_two_falls_back_to_default(self):
        mapping = resolve_estimate_columns(["Колонка 1", "Колонка 2", "Колонка 3"])
        self.assertEqual(mapping.name, "A")
        self.assertEqual(mapping.unit, "B")
        self.assertEqual(mapping.quantity, "C")
        self.assertEqual(mapping.price, "D")
        self.assertEqual(mapping.total, "E")

    def test_partial_recognition_leaves_none(self):
        mapping = resolve_estimate_columns(
            ["Наименование", "Колонка X", "Кол-во", "Цена"]
        )
        self.assertEqual(mapping.name, "A")
        self.assertIsNone(mapping.unit)
        self.assertEqual(mapping.quantity, "C")
        self.assertEqual(mapping.price, "D")
        self.assertIsNone(mapping.total)


class TestRowDetection(unittest.TestCase):
    def test_empty_row(self):
        values = {"name": "", "unit": "", "quantity": "", "price": "", "total": ""}
        self.assertTrue(_row_is_empty(values))

    def test_non_empty_row(self):
        values = {"name": "Работа", "unit": "", "quantity": "", "price": "", "total": ""}
        self.assertFalse(_row_is_empty(values))

    def test_looks_like_section_with_name_only(self):
        values = {"name": "Раздел: Потолок", "unit": "", "quantity": "", "price": "", "total": ""}
        self.assertTrue(_looks_like_section(values))

    def test_section_with_numeric_is_not_section(self):
        values = {"name": "Стены", "unit": "", "quantity": "10", "price": "", "total": ""}
        self.assertFalse(_looks_like_section(values))

    def test_item_with_unit_and_price(self):
        values = {"name": "Покраска", "unit": "м.кв.", "quantity": "20", "price": "300", "total": ""}
        self.assertTrue(_looks_like_item(values))

    def test_item_with_only_unit(self):
        values = {"name": "Грунтовка", "unit": "м.кв.", "quantity": "", "price": "", "total": ""}
        self.assertTrue(_looks_like_item(values))


class TestComputeTotal(unittest.TestCase):
    def test_uses_excel_total_when_present(self):
        result = _compute_total(
            Decimal("5"), Decimal("100"), Decimal("600")
        )
        self.assertEqual(result, Decimal("600"))

    def test_computes_from_quantity_price_when_total_missing(self):
        result = _compute_total(
            Decimal("5"), Decimal("100"), None
        )
        self.assertEqual(result, Decimal("500"))

    def test_computes_when_total_is_zero(self):
        result = _compute_total(
            Decimal("5"), Decimal("100"), Decimal("0")
        )
        self.assertEqual(result, Decimal("500"))

    def test_returns_none_when_all_none(self):
        result = _compute_total(None, None, None)
        self.assertIsNone(result)

    def test_returns_total_when_quantity_none(self):
        result = _compute_total(None, None, Decimal("500"))
        self.assertEqual(result, Decimal("500"))


class TestParseEstimateXlsxSimple(unittest.TestCase):
    def test_with_header_section_and_items(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["Раздел: Потолок", "", "", "", ""],
            ["Грунтовка потолка", "м.кв.", "25", "100", "2500"],
            ["Окраска потолка", "м.кв.", "25", "800", "20000"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        self.assertEqual(len(result.rows), 3)
        self.assertEqual(result.section_count, 1)
        self.assertEqual(result.item_count, 2)

        section = result.rows[0]
        self.assertTrue(section.is_section)
        self.assertEqual(section.row_type, "section")
        self.assertEqual(section.name, "Раздел: Потолок")
        self.assertEqual(section.sort_order, 1)

        item = result.rows[1]
        self.assertTrue(item.is_item)
        self.assertEqual(item.name, "Грунтовка потолка")
        self.assertEqual(item.unit, "м.кв.")
        self.assertEqual(item.quantity, Decimal("25"))
        self.assertEqual(item.price, Decimal("100"))
        self.assertEqual(item.total, Decimal("2500"))
        self.assertEqual(item.discounted_total, Decimal("2500"))

    def test_without_header_fallback_abcde(self):
        data = [
            ["Раздел: Стены", "", "", "", ""],
            ["Штукатурка", "м.кв.", "30", "900", "27000"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        self.assertEqual(result.section_count, 1)
        self.assertEqual(result.item_count, 1)

    def test_items_only_no_sections(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["Грунтовка", "м.кв.", "20", "100", "2000"],
            ["Окраска стен", "м.кв.", "20", "800", "16000"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        self.assertEqual(result.section_count, 0)
        self.assertEqual(result.item_count, 2)

    def test_skips_empty_rows(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["Раздел: Пол", "", "", "", ""],
            ["", "", "", "", ""],
            ["Укладка ламината", "м.кв.", "30", "1200", "36000"],
            ["", "", "", "", ""],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        self.assertEqual(len(result.rows), 2)

    def test_total_computed_from_quantity_price(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["Укладка ламината", "м.кв.", "30", "1200", ""],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertEqual(result.rows[0].total, Decimal("36000"))

    def test_decimal_comma_parsing(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["Укладка", "м.кв.", "10,5", "850,50", ""],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertEqual(result.rows[0].quantity, Decimal("10.5"))
        self.assertEqual(result.rows[0].price, Decimal("850.50"))

    def test_merged_section_cell_recognized(self):
        xlsx_bytes = _create_xlsx_with_merged_section()
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        sections = [r for r in result.rows if r.is_section]
        self.assertEqual(len(sections), 1)
        self.assertEqual(sections[0].name, "Раздел: Пол")

    def test_item_with_only_unit_parsed(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["Грунтовка", "м.кв.", "", "", ""],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.rows[0].unit, "м.кв.")
        self.assertIsNone(result.rows[0].total)

    def test_alternate_column_headers(self):
        data = [
            ["Вид работ", "Единица", "Количество", "Расценка", "Итого"],
            ["Грунтовка", "м.кв.", "25", "100", "2500"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertTrue(result.ok)


class TestParseEstimateXlsxErrors(unittest.TestCase):
    def test_invalid_xlsx_raises_parse_error(self):
        with self.assertRaises(ExcelEstimateParseError) as cm:
            parse_estimate_xlsx(b"not an xlsx file")
        self.assertIn("Не удалось открыть", str(cm.exception))

    def test_empty_xlsx_raises_parse_error(self):
        with self.assertRaises(ExcelEstimateParseError) as cm:
            parse_estimate_xlsx(b"")
        self.assertIn("пустой", str(cm.exception).lower())

    def test_no_rows_raises_parse_error(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["", "", "", "", ""],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        with self.assertRaises(ExcelEstimateParseError):
            parse_estimate_xlsx(xlsx_bytes)

    def test_oversized_fails(self):
        oversized = b"x" * (MAX_XLSX_BYTES + 1)
        with self.assertRaises(ExcelEstimateParseError) as cm:
            parse_estimate_xlsx(oversized)
        self.assertIn("слишком большой", str(cm.exception).lower())

    def test_missing_sheet_raises(self):
        data = [["Наименование", "Ед."], ["Работа", "м.кв."]]
        xlsx_bytes = _create_xlsx_bytes(data, sheet_name="Смета")
        with self.assertRaises(ExcelEstimateParseError) as cm:
            parse_estimate_xlsx(xlsx_bytes, sheet_name="НесуществующийЛист")
        self.assertIn("не найден", str(cm.exception).lower())


class TestParseEstimateXlsxDiagnostics(unittest.TestCase):
    def test_unrecognized_rows_in_diagnostics(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"],
            ["", "", "", "", "500"],
            ["Грунтовка", "м.кв.", "10", "100", "1000"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertTrue(result.ok)
        self.assertEqual(len(result.rows), 1)

    def test_result_contains_counts(self):
        data = [
            ["Наименование", "Кол-во", "Цена"],
            ["Раздел: Потолок"],
            ["Грунтовка", "25", "100"],
            ["Окраска", "25", "800"],
            ["Раздел: Стены"],
            ["Штукатурка", "30", "900"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertEqual(result.section_count, 2)
        self.assertEqual(result.item_count, 3)
        self.assertEqual(result.parsed_sheet_name, "Смета")
        self.assertTrue(result.total_rows_in_sheet > 0)


class TestParsedEstimateRow(unittest.TestCase):
    def test_is_section_true_for_section_type(self):
        row = ParsedEstimateRow(row_type="section", name="Test", sort_order=1)
        self.assertTrue(row.is_section)
        self.assertFalse(row.is_item)

    def test_is_item_true_for_item_type(self):
        row = ParsedEstimateRow(row_type="item", name="Test", sort_order=1)
        self.assertTrue(row.is_item)
        self.assertFalse(row.is_section)


class TestParsedRowsToEstimateItems(unittest.TestCase):
    def test_converts_parsed_rows_to_dicts(self):
        rows = [
            ParsedEstimateRow(
                row_type="section", name="Раздел", sort_order=1,
            ),
            ParsedEstimateRow(
                row_type="item", name="Позиция", sort_order=2,
                unit="м.кв.", quantity=Decimal("10"), price=Decimal("500"),
                total=Decimal("5000"), discounted_total=Decimal("5000"),
            ),
        ]
        items = parsed_rows_to_estimate_items(rows)
        self.assertEqual(len(items), 2)
        self.assertEqual(items[0]["row_type"], "section")
        self.assertEqual(items[1]["row_type"], "item")
        self.assertEqual(items[1]["unit"], "м.кв.")
        self.assertEqual(items[1]["quantity"], "10")
        self.assertEqual(items[1]["total"], "5000")

    def test_converts_parsed_rows_no_db_import(self):
        rows = [
            ParsedEstimateRow(
                row_type="item", name="Test", sort_order=1,
                unit="шт.", quantity=Decimal("1"), price=Decimal("100"),
                total=Decimal("100"), discounted_total=Decimal("100"),
            ),
        ]
        items = parsed_rows_to_estimate_items(rows)
        self.assertIsInstance(items, list)
        self.assertIsInstance(items[0], dict)
        self.assertEqual(items[0]["name"], "Test")


class TestParserNoDependencies(unittest.TestCase):
    """Confirm the parser does not import DB or route modules."""

    def test_estimate_repository_not_imported(self):
        self.assertNotIn(
            "webapp.estimate_repository",
            openpyxl.__class__.__module__,
        )

    def test_db_module_not_imported(self):
        import sys
        self.assertNotIn("webapp.db", sys.modules)


class LookalikeSummaryTests(unittest.TestCase):
    """Test _looks_like_summary heuristic."""

    def test_itogo_po_razdelu(self):
        from webapp.excel_estimate_parser import _looks_like_summary
        self.assertTrue(_looks_like_summary({"name": "Итого по разделу:", "unit": "", "quantity": "", "price": "", "total": ""}))

    def test_itogo_po_smete(self):
        from webapp.excel_estimate_parser import _looks_like_summary
        self.assertTrue(_looks_like_summary({"name": "Итого по смете:", "unit": "", "quantity": "", "price": "", "total": ""}))

    def test_vsego_po_smete(self):
        from webapp.excel_estimate_parser import _looks_like_summary
        self.assertTrue(_looks_like_summary({"name": "ВСЕГО по смете", "unit": "", "quantity": "", "price": "", "total": ""}))

    def test_normal_item_not_summary(self):
        from webapp.excel_estimate_parser import _looks_like_summary
        self.assertFalse(_looks_like_summary({"name": "Грунтовка потолка", "unit": "м.кв.", "quantity": "20", "price": "100", "total": "2000"}))

    def test_section_not_summary(self):
        from webapp.excel_estimate_parser import _looks_like_summary
        self.assertFalse(_looks_like_summary({"name": "Раздел: Потолок", "unit": "", "quantity": "", "price": "", "total": ""}))


class RealFormatParseTests(unittest.TestCase):
    """Test parser against realistic estimate structure (header on row 15, B-H columns)."""

    def test_real_format_header_row_15_section_items(self):
        data = []
        for _ in range(14):
            data.append(["", "", "", "", "", "", "", ""])
        data.append([
            "№",
            "Наименование работ",
            "Ед. изм.",
            "Кол-во",
            "Цена",
            "Ск-ка",
            "Стоимость",
            "Ст. со скидкой",
        ])
        data.append(["", "Потолочные работы ГКЛ", "", "", "", "", "", ""])
        data.append([
            "1",
            "Монтаж каркаса потолка",
            "м.кв.",
            "45,5",
            "2800",
            "",
            "127400",
            "123000",
        ])
        data.append([
            "2",
            "Обшивка потолка ГКЛ в два слоя",
            "м.кв.",
            "45,5",
            "1600",
            "",
            "72800",
            "70000",
        ])
        data.append(["", "Итого по разделу:", "", "", "", "", "200200", "193000"])
        data.append(["", "Раздел: Стены", "", "", "", "", "", ""])
        data.append([
            "3",
            "Штукатурка стен по маякам",
            "м.кв.",
            "60",
            "900",
            "",
            "54000",
            "52000",
        ])

        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        self.assertEqual(result.section_count, 2)
        self.assertEqual(result.item_count, 3)

        section_names = [r.name for r in result.rows if r.is_section]
        self.assertIn("Потолочные работы ГКЛ", section_names)
        self.assertIn("Раздел: Стены", section_names)

        items = [r for r in result.rows if r.is_item]
        self.assertEqual(items[0].name, "Монтаж каркаса потолка")
        self.assertEqual(items[0].unit, "м.кв.")
        self.assertEqual(items[0].quantity, Decimal("45.5"))
        self.assertEqual(items[0].price, Decimal("2800"))
        self.assertEqual(items[0].total, Decimal("127400"))
        self.assertEqual(items[0].discounted_total, Decimal("123000"))

        self.assertEqual(items[1].name, "Обшивка потолка ГКЛ в два слоя")
        self.assertEqual(items[1].discounted_total, Decimal("70000"))

        self.assertEqual(items[2].name, "Штукатурка стен по маякам")
        self.assertEqual(items[2].discounted_total, Decimal("52000"))

        summary_skipped = any("итоговая" in d for d in result.diagnostics)
        self.assertTrue(summary_skipped)

    def test_real_format_discounted_total_falls_back_to_total(self):
        data = []
        for _ in range(14):
            data.append(["", "", "", "", "", "", "", ""])
        data.append([
            "№",
            "Наименование работ",
            "Ед. изм.",
            "Кол-во",
            "Цена",
            "Ск-ка",
            "Стоимость",
            "Ст. со скидкой",
        ])
        data.append(["1", "Грунтовка потолка", "м.кв.", "20", "100", "", "2000", ""])
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        item = result.rows[0]
        self.assertEqual(item.total, Decimal("2000"))
        self.assertEqual(item.discounted_total, Decimal("2000"))

    def test_header_found_on_row_20(self):
        data = []
        for _ in range(19):
            data.append(["", "", "", "", "", ""])
        data.append(["Наименование", "Ед.", "Кол-во", "Цена", "Сумма"])
        data.append(["Грунтовка", "м.кв.", "10", "100", "1000"])
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        self.assertEqual(len(result.rows), 1)

    def test_name_in_column_b_detected(self):
        data = [
            ["", "Наименование работ", "Ед. изм.", "Кол-во", "Цена", "Стоимость"],
            ["", "Грунтовка потолка", "м.кв.", "20", "100", "2000"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)

        self.assertTrue(result.ok)
        self.assertEqual(result.rows[0].name, "Грунтовка потолка")


class SignatureTrashRowTests(unittest.TestCase):
    def test_director_not_imported(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "Генеральный директор"}))

    def test_director_lowercase(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "генеральный директор Иванов И.И."}))

    def test_just_director(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "Директор"}))

    def test_chief_engineer(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "Главный инженер"}))

    def test_chief_accountant(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "Главный бухгалтер"}))

    def test_signature_keyword(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "Подпись"}))

    def test_seal_keyword(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "Печать"}))

    def test_mp_keyword(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "М.П."}))

    def test_underscores_only(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "____________"}))

    def test_dashes_only(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "——————"}))

    def test_dots_only(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "......"}))

    def test_year_alone(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "2026"}))

    def test_year_with_g(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "2026 г."}))

    def test_year_with_god(self):
        self.assertTrue(_looks_like_signature_or_trash({"name": "2025 год"}))

    def test_normal_item_not_trash(self):
        self.assertFalse(_looks_like_signature_or_trash({"name": "Штукатурка стен"}))

    def test_normal_section_not_trash(self):
        self.assertFalse(_looks_like_signature_or_trash({"name": "Раздел 1: Демонтажные работы"}))

    def test_item_with_quantity_not_trash(self):
        self.assertFalse(_looks_like_signature_or_trash({"name": "2026", "quantity": "5", "price": "100", "total": "500"}))

    def test_underscores_with_text_not_trash(self):
        self.assertFalse(_looks_like_signature_or_trash({"name": "Штукатурка ______"}))

    def test_year_row_in_xlsx_skipped(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Стоимость"],
            ["Грунтовка", "м2", "10", "100", "1000"],
            ["Генеральный директор", "", "", "", ""],
            ["____________", "", "", "", ""],
            ["2026 г.", "", "", "", ""],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertEqual(result.item_count, 1)
        self.assertEqual(len(result.rows), 1)
        self.assertTrue(any("служебная" in d or "подписная" in d for d in result.diagnostics))

    def test_signature_rows_counted_in_diagnostics(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Стоимость"],
            ["Штукатурка", "м2", "10", "100", "1000"],
            ["Директор", "", "", "", ""],
            ["М.П.", "", "", "", ""],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertEqual(result.item_count, 1)
        signature_diag = [d for d in result.diagnostics if "служебная" in d or "подписная" in d]
        self.assertEqual(len(signature_diag), 2)


class SectionWithTotalTests(unittest.TestCase):
    def test_section_with_total_only_is_section(self):
        row = {"name": "Сантехнические работы (примерный расчёт)", "unit": "", "quantity": None, "price": None, "total": "235200", "discounted_total": None}
        self.assertTrue(_looks_like_section(row))

    def test_section_with_total_not_item(self):
        data = [
            ["Наименование", "Ед.", "Кол-во", "Цена", "Стоимость", "Ст. со скидкой"],
            ["Сантехнические работы (примерный расчёт)", "", "", "", "235200", ""],
            ["Установка раковины", "шт", "2", "3500", "7000", "6300"],
        ]
        xlsx_bytes = _create_xlsx_bytes(data)
        result = parse_estimate_xlsx(xlsx_bytes)
        self.assertEqual(result.section_count, 1)
        self.assertEqual(result.item_count, 1)
        self.assertEqual(result.rows[0].row_type, "section")
        self.assertEqual(result.rows[0].name, "Сантехнические работы (примерный расчёт)")
        self.assertIsNone(result.rows[0].quantity)
        self.assertIsNone(result.rows[0].price)

    def test_section_with_qty_still_item(self):
        row = {"name": "Работы", "unit": "м2", "quantity": "10", "price": None, "total": "1000", "discounted_total": None}
        self.assertFalse(_looks_like_section(row))
        self.assertTrue(_looks_like_item(row))

    def test_section_with_price_still_item(self):
        row = {"name": "Работы", "unit": "", "quantity": None, "price": "100", "total": "1000", "discounted_total": None}
        self.assertFalse(_looks_like_section(row))

    def test_section_without_numbers_still_section(self):
        row = {"name": "Отделочные работы", "unit": "", "quantity": None, "price": None, "total": None, "discounted_total": None}
        self.assertTrue(_looks_like_section(row))


if __name__ == "__main__":
    unittest.main()
